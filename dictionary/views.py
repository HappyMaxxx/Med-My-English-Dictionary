
from django.http import HttpResponseRedirect, HttpResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.urls import reverse_lazy, reverse
import openpyxl
from .forms import AddWordForm, WordForm, GroupForm
from django.views.generic import ListView, CreateView
from django.views import View
import requests

from reportlab.pdfbase.ttfonts import TTFont
from reportlab.pdfbase import pdfmetrics
from reportlab.lib.pagesizes import letter
from reportlab.pdfgen import canvas

from django.contrib.auth.decorators import login_required
from django.db.models import Q

from django.http import JsonResponse
import json

from django.contrib.auth.mixins import LoginRequiredMixin
from django.conf import settings
import os

from celery import shared_task, group
from celery.result import AsyncResult

from django.contrib.auth.models import User
from med.models import UserProfile
from .models import Word, WordGroup

from med.views import add_to_main_group
from notifications.views import create_notification

import logging

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

MAX_GROUP_COUNT = 20
font_path = os.path.join(settings.BASE_DIR, 'dictionary/static/dictionary/fonts/DejaVuSans.ttf')
bold_font_path = os.path.join(settings.BASE_DIR, 'dictionary/static/dictionary/fonts/DejaVuSans-Bold.ttf')

pdfmetrics.registerFont(TTFont('DejaVuSans', font_path))
pdfmetrics.registerFont(TTFont('DejaVuSans-Bold', bold_font_path))

class AddWordView(LoginRequiredMixin, CreateView):
    form_class = AddWordForm
    template_name = 'dictionary/add_word.html'
    extra_context = {'title': 'Add Word'}

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['from_text'] = self.request.GET.get('from') == 'text'
        return context

    def get_success_url(self):
        return reverse_lazy('words', kwargs={'user_name': self.request.user.username})

    def form_valid(self, form):
        word = form.save(commit=False)
        word.user = self.request.user
        word.save()

        add_to_main_group(self.request, word)

        from_text = self.request.POST.get('from_text', '') == 'true'
        if from_text:
            user_profile = UserProfile.objects.get(user=self.request.user)
            user_profile.words_added_from_text += 1
            user_profile.save()

        return super().form_valid(form)
    

class EditWordView(LoginRequiredMixin, View):
    def get(self, request, word_id, *args, **kwargs):
        word = get_object_or_404(Word, id=word_id, user=request.user)
        form = WordForm(instance=word)
        return render(request, 'dictionary/edit_word.html', {'form': form, 'word': word})

    def post(self, request, word_id, *args, **kwargs):
        word = get_object_or_404(Word, id=word_id, user=request.user)
        form = WordForm(request.POST, instance=word)
        if form.is_valid():
            form.save()

            if form.has_changed():
                user_profile = UserProfile.objects.get(user=request.user)
                user_profile.edited_words += 1
                user_profile.save()

            return redirect('words', user_name=request.user.username)
        return render(request, 'dictionary/edit_word.html', {'form': form, 'word': word})
    

class WordListView(LoginRequiredMixin, ListView):
    model = Word
    template_name = 'dictionary/words.html'
    context_object_name = 'words'
    paginate_by = 25

    def get(self, request, *args, **kwargs):
        if 'sort_alphabet' not in request.GET and 'sort_date' not in request.GET:
            query_params = request.GET.copy()
            query_params['sort_date'] = 'desc' 
            return HttpResponseRedirect(f"{request.path}?{query_params.urlencode()}")
        return super().get(request, *args, **kwargs)

    def get_queryset(self):
        user_name = self.kwargs['user_name']
        user = get_object_or_404(User, username=user_name)

        self.is_my_dict = user == self.request.user
        queryset = Word.objects.filter(user=user)
        
        word_type = self.request.GET.get('type')
        if word_type:
            queryset = queryset.filter(word_type=word_type)

        filter_word = self.request.GET.get('filter_word', '')
        filter_translation = self.request.GET.get('filter_translation', '')

        if filter_word:
            queryset = queryset.filter(word__icontains=filter_word)
        if filter_translation:
            queryset = queryset.filter(translation__icontains=filter_translation)

        sort_alphabet = self.request.GET.get('sort_alphabet')
        sort_date = self.request.GET.get('sort_date')

        if sort_alphabet == 'asc':
            queryset = queryset.order_by('word')
        elif sort_alphabet == 'desc':
            queryset = queryset.order_by('-word')

        if sort_date == 'asc':
            queryset = queryset.order_by('time_create')
        elif sort_date == 'desc':
            queryset = queryset.order_by('-time_create')

        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        user = get_object_or_404(User, username=self.kwargs['user_name'])
        request_user = self.request.user

        friends = User.objects.filter(
            Q(friendship_requests_sent__receiver=user, friendship_requests_sent__status='accepted') |
            Q(friendship_requests_received__sender=user, friendship_requests_received__status='accepted')
        ).distinct()

        is_friends = request_user in friends

        types = Word.TYPE_CHOICES

        context.update({
            'user_name': self.kwargs['user_name'],
            'user': user,
            'title': f"{self.kwargs['user_name']}'s Dictionary",
            'is_my_dict': getattr(self, 'is_my_dict', False),
            'is_dict': True,
            'logged_user': request_user,
            'access': UserProfile.objects.get(user=user).access_dictionary,
            'is_friends': is_friends,
            'filter_word': self.request.GET.get('filter_word', ''),
            'filter_translation': self.request.GET.get('filter_translation', ''),
            'sort_alphabet': self.request.GET.get('sort_alphabet', 'asc'),
            'sort_date': self.request.GET.get('sort_date', 'asc'),
            'types': types,
        })
        return context

def make_favourite(request, word_id):
    try:
        word = get_object_or_404(Word, id=word_id, user=request.user)
        word.is_favourite = not word.is_favourite
        word.save()
        return redirect('words', user_name=request.user.username)
    except:
        return redirect('login')

def save_all_words_as_json(request):
    user = request.user
    words = Word.objects.filter(user=user)
    data = []

    for word in words:
        data.append({
            'word': word.word,
            'translation': word.translation,
            'example': word.example,
            'word_type': word.word_type,
        })
    
    file_name = f"{user.username}_words.json"
    file_path = os.path.join('media', file_name)

    with open(file_path, 'w') as file:
        json.dump(data, file, indent=4)

    return redirect('download_file', file=file_name)

def find_word_type(request):
    if request.method == 'POST':
        data = json.loads(request.body)
        word = data.get('word', None)

        if word:
            api_url = f"https://api.dictionaryapi.dev/api/v2/entries/en/{word}"
            response = requests.get(api_url)

            if response.status_code == 200:
                word_type = response.json()[0]['meanings'][0]['partOfSpeech']
                try:
                    all_types = set()
                    for meaning in response.json()[0]['meanings']:
                        all_types.add(meaning['partOfSpeech'])
                    all_types.remove(word_type)
                    all_types = list(all_types)

                    if len(all_types) > 1:
                        return JsonResponse({'word_type': word_type, 'all_types': all_types}, status=200)
                    else:
                        return JsonResponse({'word_type': word_type}, status=200)
                except:
                    return JsonResponse({'word_type': word_type}, status=200)
            else:
                return JsonResponse({'word_type': 'other'}, status=200)
        
        return JsonResponse({'word_type': 'other'}, status=200)
    
    return JsonResponse({'error': 'Invalid request method'}, status=400)

def check_word(request):
    if request.method == 'POST':
        data = json.loads(request.body)
        word = data.get('word', None)
        if word:
            if Word.objects.filter(word=word, user=request.user).exists():
                return JsonResponse({'error': 'This word already exists in your dictionary.'}, status=200)
            return JsonResponse({'error': ''}, status=200)
        return JsonResponse({'error': 'Invalid word'}, status=200)

def get_word_type(word):
    if not word:
        return 'other'
    api_url = f"https://api.dictionaryapi.dev/api/v2/entries/en/{word}"
    response = requests.get(api_url)
    if response.status_code == 200:
        try:
            word_type = response.json()[0]['meanings'][0]['partOfSpeech']
            return word_type
        except:
            return 'other'
    return 'other'

@shared_task
def process_word(word_data, user_id, group_id):
    word_type = get_word_type(word_data['word'])
    word = Word.objects.create(
        word=word_data['word'],
        translation=word_data['translation'],
        example=word_data['example'],
        user_id=user_id,
        word_type=word_type,
    )
    word_group = WordGroup.objects.get(id=group_id)
    word_group.words.add(word)
    return word.id

@shared_task(bind=True, max_retries=10)
def notify_file_processed(self, user_id, file_name, task_ids):
    all_done = True
    for task_id in task_ids:
        result = AsyncResult(task_id)
        if not result.ready():
            all_done = False
            break
        if not result.successful():
            create_notification(
                receiver=User.objects.get(pk=user_id),
                message=f"Some words from file '{file_name}' failed to process."
            )
            return
    if not all_done:
        self.retry(countdown=10)
    create_notification(
        receiver=User.objects.get(pk=user_id),
        message=f"All words from file '{file_name}' have been successfully added."
    )

def upload_file(request):
    if request.method == 'POST' and request.FILES.get('file'):
        group_name = f"All {request.user.username}'s "
        word_group, created = WordGroup.objects.get_or_create(
            name=group_name,
            is_main=True,
            user=request.user
        )
        uploaded_file = request.FILES['file']
        file_name = uploaded_file.name.lower()
        try:
            if file_name.endswith('.xlsx') or file_name.endswith('.xls'):
                wb = openpyxl.load_workbook(uploaded_file)
                sheet = wb[wb.sheetnames[0]]
                words = []
                for row in sheet.iter_rows(values_only=True):
                    if len(row) == 3:
                        words.append(row)
                wb.close()
                if words[0] == ('Word', 'Translation', 'Example'):
                    tasks = [
                        process_word.s({
                            'word': word_data[0],
                            'translation': word_data[1],
                            'example': word_data[2]
                        }, request.user.id, word_group.id)
                        for word_data in words[1:]
                    ]
                    if tasks:
                        result = group(tasks).apply_async()
                        task_ids = [task.id for task in result.children]
                        notify_file_processed.delay(request.user.id, uploaded_file.name, task_ids)
                    create_notification(
                        receiver=request.user,
                        message=f"Processing of file '{uploaded_file.name}' started. Words will be added soon."
                    )
                    return redirect('words', user_name=request.user.username)
                else:
                    create_notification(
                        receiver=request.user,
                        message="Excel file format is incorrect. First row should be 'Word', 'Translation', 'Example'."
                    )
                    return render(request, 'dictionary/words_ff.html')
            elif file_name.endswith('.txt'):
                content = uploaded_file.read().decode('utf-8')
                try:
                    json_data = json.loads(content)
                    if isinstance(json_data, list):
                        tasks = []
                        for item in json_data:
                            if all(k in item for k in ('word', 'translation', 'example')):
                                word_type = item.get('word_type')
                                if not word_type:
                                    tasks.append(process_word.s({
                                        'word': item['word'],
                                        'translation': item['translation'],
                                        'example': item['example']
                                    }, request.user.id, word_group.id))
                                else:
                                    word = Word.objects.create(
                                        word=item['word'],
                                        translation=item['translation'],
                                        example=item['example'],
                                        word_type=word_type,
                                        user=request.user,
                                    )
                                    word_group.words.add(word)
                        if tasks:
                            result = group(tasks).apply_async()
                            task_ids = [task.id for task in result.children]
                            notify_file_processed.delay(request.user.id, uploaded_file.name, task_ids)
                        create_notification(
                            receiver=request.user,
                            message=f"Processing of file '{uploaded_file.name}' started. Words will be added soon."
                        )
                        return redirect('words', user_name=request.user.username)
                    else:
                        create_notification(
                            receiver=request.user,
                            message="JSON file should contain a list of objects with keys: 'word', 'translation', 'example'."
                        )
                        return render(request, 'dictionary/words_ff.html')
                except json.JSONDecodeError:
                    create_notification(
                        receiver=request.user,
                        message="Invalid JSON format in text file."
                    )
                    return render(request, 'dictionary/words_ff.html')
            else:
                create_notification(
                    receiver=request.user,
                    message="Unsupported file format. Please upload .xlsx, .xls, or .txt file."
                )
                return render(request, 'dictionary/words_ff.html')
        except Exception as e:
            create_notification(
                receiver=request.user,
                message=f"An error occurred while processing the file: {e}"
            )
            return render(request, 'dictionary/words_ff.html')
    return render(request, 'dictionary/words_ff.html')

def wrap_text(text, max_width, font, font_size, pdf):
    words = text.split()
    lines = []
    line = []

    for word in words:
        line.append(word)
        if pdf.stringWidth(' '.join(line), font, font_size) > max_width:
            line.pop()
            lines.append(' '.join(line))
            line = [word]

    if line:
        lines.append(' '.join(line))

    return lines

@login_required
def export_pdf(request):

    if not request.user.user_profile.is_premium:
        return redirect('soon')
    
    words = Word.objects.filter(user=request.user)
    
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{request.user.username}_dictionary.pdf"'
    
    pdf = canvas.Canvas(response, pagesize=letter)
    width, height = letter
    left_margin = 50
    right_margin = 50
    max_width = width - left_margin - right_margin
    y = height - 50

    pdf.setFont("DejaVuSans", 12)
    pdf.drawString(left_margin, y, f"Dictionary of {request.user.username}")
    y -= 30 
    pdf.setFont("DejaVuSans", 10)
    pdf.drawString(left_margin, y, "-" * 50)
    y -= 20

    for word in words:
        pdf.setFont("DejaVuSans-Bold", 10)
        pdf.drawString(left_margin, y, f"Word: {word.word}")
        y -= 15

        pdf.setFont("DejaVuSans", 10)
        translation_lines = wrap_text(f"Translation: {word.translation}", max_width, "DejaVuSans", 10, pdf)
        for line in translation_lines:
            pdf.drawString(left_margin, y, line)
            y -= 15

        example_lines = wrap_text(f"Example: {word.example}", max_width, "DejaVuSans", 10, pdf)
        for line in example_lines:
            pdf.drawString(left_margin, y, line)
            y -= 15

        y -= 10

        if y < 50:
            pdf.showPage()
            pdf.setFont("DejaVuSans", 10)
            y = height - 100

    pdf.save()
    return response


# GROUPS LOGIC
class BaseGroupView(ListView):
    template_name = 'med/groups.html'

    def get_common_context(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['groups'] = WordGroup.objects.filter(user=self.request.user).order_by('-is_main', 'name')
        context['used_groups'] = WordGroup.objects.filter(uses_users=self.request.user)
        context['len_groups'] = max(context['groups'].count() + context['used_groups'].count() - 1, 0)
        context['max_group_count'] = MAX_GROUP_COUNT if not self.request.user.user_profile.grouper else 100
        return context


class GroupListView(BaseGroupView):
    model = WordGroup

    def get_context_data(self, **kwargs):
        context = self.get_common_context(**kwargs)
        context['title'] = "Groups"
        context['title1'] = "Words"
        context['is_group'] = False
        return context


class GroupWordsView(BaseGroupView):
    model = Word
    context_object_name = 'words'
    paginate_by = 25

    def dispatch(self, request, *args, **kwargs):
        group_id = self.kwargs.get('group_id')
        group = get_object_or_404(WordGroup, id=group_id)

        if request.user != group.user and request.user not in group.uses_users.all():
            return HttpResponseRedirect(reverse('groups'))
        
        return super().dispatch(request, *args, **kwargs)

    def get_context_data(self, **kwargs):
        context = self.get_common_context(**kwargs)

        group_id = self.kwargs.get('group_id')
        group = get_object_or_404(WordGroup, id=group_id)
        is_my_group = group.user == self.request.user

        friends = User.objects.filter(
            Q(friendship_requests_sent__receiver=self.request.user, friendship_requests_sent__status='accepted') |
            Q(friendship_requests_received__sender=self.request.user, friendship_requests_received__status='accepted')
        ).distinct()

        context.update({
            'is_my_group': is_my_group,
            'is_uses': not is_my_group and group.uses_users.filter(id=self.request.user.id).exists(),
            'title': "Groups",
            'title1': f"{group.name} Words ({group.user.username})" if not is_my_group else f"{group.name} Words",
            'is_main': group.is_main,
            'group_id': group_id,
            'group_': group,
            'is_group': True,
            'words_f_g': True,
            'friends': friends,
        })

        if context['is_uses']:
            all_words = Word.objects.filter(user=self.request.user)

            user_word_titles = [word.word.lower() for word in all_words]

            group_words = group.words.all()

            for word in group_words:
                word.is_saved = word.word.lower() in user_word_titles

            context['words'] = group_words
        return context

    def get_queryset(self):
        group_id = self.kwargs.get('group_id')
        group = get_object_or_404(WordGroup, id=group_id)
        return group.words.all()


class CreateGroupView(LoginRequiredMixin, View):
    def get(self, request, *args, **kwargs):

        len_groups = WordGroup.objects.filter(user=request.user).count()

        if request.user.user_profile.grouper and len_groups >= 101:
            return redirect('groups')

        if len_groups >= (MAX_GROUP_COUNT + 1) and not request.user.user_profile.grouper:
            return redirect('groups')
        
        form = GroupForm()
        return render(request, 'med/create_group.html', {'form': form})

    def post(self, request, *args, **kwargs):
        form = GroupForm(request.POST)
        if form.is_valid():
            group_name = form.cleaned_data['name']
            if WordGroup.objects.filter(name=group_name, user=request.user).exists():
                form.add_error('name', "Group with this name already exists")
                return render(request, 'med/create_group.html', {'form': form})

            group = form.save(commit=False)
            group.user = request.user
            group.save()
            return redirect('groups')

        return render(request, 'med/create_group.html', {'form': form})
    

class SelectGroupView(View):
    def get(self, request):
        word_ids = request.GET.getlist('word_ids')
        if not word_ids:
            return redirect('words', user_name=request.user.username)
        words = Word.objects.filter(id__in=word_ids)
        groups = WordGroup.objects.filter(user=request.user, is_main=False)
        return render(request, 'med/select_group.html', {
            'words': words,
            'word_ids': word_ids,
            'groups': groups,
        })

    def post(self, request, *args, **kwargs):
        word_ids = request.POST.getlist('word_ids')
        group_id = request.POST.get('group')

        if group_id:
            group = get_object_or_404(WordGroup, id=group_id, user=request.user)
        else:
            redirect('words', user_name=request.user.username)
        
        group_words = group.words.all()
        words = Word.objects.filter(id__in=word_ids, user=request.user)

        for word in words:
            if word not in group_words:
                group.words.add(word)

        group.save()

        return redirect('group_words', group_id=group_id)