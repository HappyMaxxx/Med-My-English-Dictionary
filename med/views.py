from datetime import datetime
import re
import string
from django.http import FileResponse, Http404, HttpResponseNotFound, HttpResponseRedirect, HttpResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.urls import reverse, reverse_lazy
from med.forms import AddWordForm, ChengePasswordForm, RegisterUserForm, LoginUserForm, WordForm, GroupForm, EditProfileForm, AvatarUpdateForm, WordsShowForm, TextForm
from django.views.generic import ListView, CreateView
from django.contrib.auth.views import LoginView
from django.views import View
from django.views.decorators.cache import cache_page
from django.db.models import Min, Max, When, Case
import requests

from reportlab.pdfbase.ttfonts import TTFont
from reportlab.pdfbase import pdfmetrics
from reportlab.lib.pagesizes import letter
from reportlab.pdfgen import canvas

from django.utils.timezone import now, timedelta
from django.db.models.functions import TruncDay

from django.core.files.base import ContentFile
import base64

from django.contrib.auth.decorators import login_required
from django.contrib import messages

from django.db import transaction
from django.db.models import Q, Count

from django.http import JsonResponse
import json

from django.core.paginator import Paginator

from urllib.parse import urlparse

import openpyxl

from django.contrib.auth.tokens import default_token_generator
from django.utils.http import urlsafe_base64_decode
from django.contrib.auth import logout, login, update_session_auth_hash
from django.contrib.auth.mixins import LoginRequiredMixin
import os
from med.models import *

from django.db.models.signals import post_save, m2m_changed
from django.dispatch import receiver
from django.conf import settings

from med.tasks import send_activation_email, update_top

MAX_GROUP_COUNT = 20

practice_cards = {
    'test': {
        'word': 'Test',
        'img': 'med/img/practice/dice_light.png',
        'dark_img': 'med/img/practice/dice_dark.png',
        'href': 'soon'
    },
    'reading': {
        'word': 'Reading',
        'img': 'med/img/practice/read_light.png',
        'dark_img': 'med/img/practice/read_dark.png',
        'href': 'practice_reading'
    },
    'groups': {
        'word': 'Groups',
        'img': 'med/img/practice/group_light.png',
        'dark_img': 'med/img/practice/group_dark.png',
        'href': 'practice_groups'
    },
}

processed_signals = {}
process_wrods_signals = {}

@cache_page(60 * 15)
def index(request):
    if request.user.is_authenticated:
        return redirect('profile', user_name=request.user.username)
    return render(request, 'med/index.html')

@cache_page(60 * 15)
def about(request):
    return render(request, 'med/about.html')

def add_to_main_group(request, word):
    group_name = f"All {request.user.username}'s "
    group, created = WordGroup.objects.get_or_create(
        name=group_name,
        is_main=True,
        user=request.user
    )

    group.save()

    if isinstance(word, str):
        word = Word.objects.get(word=word, user=request.user)

    group.words.add(word)


class AddWordView(LoginRequiredMixin, CreateView):
    form_class = AddWordForm
    template_name = 'med/add_word.html'
    extra_context = {'title': 'Add Word'}

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['from_text'] = self.request.GET.get('from') == 'text'
        return context

    def get_success_url(self):
        return reverse_lazy('words', kwargs={'user_name': self.request.user.username})

    def form_valid(self, form):
        word = form.save(commit=False)
        word.user = self.request.user
        word.save()

        add_to_main_group(self.request, word)

        from_text = self.request.POST.get('from_text', '') == 'true'
        if from_text:
            user_profile = UserProfile.objects.get(user=self.request.user)
            user_profile.words_added_from_text += 1
            user_profile.save()

        return super().form_valid(form)


class ConfirmDeleteView(LoginRequiredMixin, View):
    def get(self, request, *args, **kwargs):
        word_ids = request.GET.getlist('word_ids')
        group_id = request.GET.get('group_id')
        text_id = request.GET.get('text_id')

        context = {}
        if text_id:
            text = get_object_or_404(ReadingText, id=text_id)
            context.update({
                'is_text': True,
                'text': 'Are you sure you want to delete this text?',
                'text_id': text_id,
                'text_title': text.title
            })

        elif group_id:
            group = get_object_or_404(WordGroup, id=group_id, user=request.user)
            words = group.words.filter(id__in=word_ids) if word_ids else []
            context.update({
                'is_group': True,
                'group_name': group.name,
                'group_id': group_id,
                'words': words,
                'text': self._get_delete_message(words, group.name)
            })

            if word_ids:
                context['word_ids'] = word_ids

        elif word_ids:
            words = Word.objects.filter(id__in=word_ids, user=request.user)
            if not words:
                return redirect('words')
            context.update({
                'is_group': False,
                'words': words,
                'word_ids': word_ids,
                'text': self._get_delete_message(words)
            })
        return render(request, 'med/confirm_delete.html', context)

    def post(self, request, *args, **kwargs):
        word_ids = request.POST.getlist('word_ids')
        group_id = request.POST.get('group_id')
        text_id = request.POST.get('text_id')

        if text_id:
            get_object_or_404(ReadingText, id=text_id).delete()
            return redirect('practice_reading')
        elif group_id:
            group = get_object_or_404(WordGroup, id=group_id, user=request.user)
            if word_ids:
                words = group.words.filter(id__in=word_ids)
                for word in words:
                    group.words.remove(word)

                return redirect('group_words', group_id=group_id)
            else:
                group.delete()
                return redirect('groups')
        elif word_ids:
            Word.objects.filter(id__in=word_ids, user=request.user).delete()
            return redirect('words', user_name=request.user.username)
        
        return redirect('profile', user_name=request.user.username)

    @staticmethod
    def _get_delete_message(words, group_name=None):
        if group_name:
            return ("Are you sure you want to delete this group?"
                    if len(words) == 0 else
                    f"Are you sure you want to delete these words from group {group_name}?"
                    if len(words) > 1 else
                    f"Are you sure you want to delete this word from group {group_name}?")
        return ("Are you sure you want to delete these words?"
                if len(words) > 1 else
                "Are you sure you want to delete this word?")


class EditWordView(LoginRequiredMixin, View):
    def get(self, request, word_id, *args, **kwargs):
        word = get_object_or_404(Word, id=word_id, user=request.user)
        form = WordForm(instance=word)
        return render(request, 'med/edit_word.html', {'form': form, 'word': word})

    def post(self, request, word_id, *args, **kwargs):
        word = get_object_or_404(Word, id=word_id, user=request.user)
        form = WordForm(request.POST, instance=word)
        if form.is_valid():
            form.save()

            if form.has_changed():
                user_profile = UserProfile.objects.get(user=request.user)
                user_profile.edited_words += 1
                user_profile.save()

            return redirect('words', user_name=request.user.username)
        return render(request, 'med/edit_word.html', {'form': form, 'word': word})
    

class WordListView(LoginRequiredMixin, ListView):
    model = Word
    template_name = 'med/words.html'
    context_object_name = 'words'
    paginate_by = 25

    def get(self, request, *args, **kwargs):
        if 'sort_alphabet' not in request.GET and 'sort_date' not in request.GET:
            query_params = request.GET.copy()
            query_params['sort_date'] = 'desc' 
            return HttpResponseRedirect(f"{request.path}?{query_params.urlencode()}")
        return super().get(request, *args, **kwargs)

    def get_queryset(self):
        user_name = self.kwargs['user_name']
        user = get_object_or_404(User, username=user_name)

        self.is_my_dict = user == self.request.user
        queryset = Word.objects.filter(user=user)
        
        word_type = self.request.GET.get('type')
        if word_type:
            queryset = queryset.filter(word_type=word_type)

        filter_word = self.request.GET.get('filter_word', '')
        filter_translation = self.request.GET.get('filter_translation', '')

        if filter_word:
            queryset = queryset.filter(word__icontains=filter_word)
        if filter_translation:
            queryset = queryset.filter(translation__icontains=filter_translation)

        sort_alphabet = self.request.GET.get('sort_alphabet')
        sort_date = self.request.GET.get('sort_date')

        if sort_alphabet == 'asc':
            queryset = queryset.order_by('word')
        elif sort_alphabet == 'desc':
            queryset = queryset.order_by('-word')

        if sort_date == 'asc':
            queryset = queryset.order_by('time_create')
        elif sort_date == 'desc':
            queryset = queryset.order_by('-time_create')

        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        user = get_object_or_404(User, username=self.kwargs['user_name'])
        request_user = self.request.user

        friends = User.objects.filter(
            Q(friendship_requests_sent__receiver=user, friendship_requests_sent__status='accepted') |
            Q(friendship_requests_received__sender=user, friendship_requests_received__status='accepted')
        ).distinct()

        is_friends = request_user in friends

        types = Word.TYPE_CHOICES

        context.update({
            'user_name': self.kwargs['user_name'],
            'user': user,
            'title': f"{self.kwargs['user_name']}'s Dictionary",
            'is_my_dict': getattr(self, 'is_my_dict', False),
            'is_dict': True,
            'logged_user': request_user,
            'access': UserProfile.objects.get(user=user).access_dictionary,
            'is_friends': is_friends,
            'filter_word': self.request.GET.get('filter_word', ''),
            'filter_translation': self.request.GET.get('filter_translation', ''),
            'sort_alphabet': self.request.GET.get('sort_alphabet', 'asc'),
            'sort_date': self.request.GET.get('sort_date', 'asc'),
            'types': types,
        })
        return context


class BaseGroupView(ListView):
    template_name = 'med/groups.html'

    def get_common_context(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['groups'] = WordGroup.objects.filter(user=self.request.user).order_by('-is_main', 'name')
        context['used_groups'] = WordGroup.objects.filter(uses_users=self.request.user)
        context['len_groups'] = max(context['groups'].count() + context['used_groups'].count() - 1, 0)
        context['max_group_count'] = MAX_GROUP_COUNT if self.request.user.username != 'grouper' else 100
        return context


class GroupListView(BaseGroupView):
    model = WordGroup

    def get_context_data(self, **kwargs):
        context = self.get_common_context(**kwargs)
        context['title'] = "Groups"
        context['title1'] = "Words"
        context['is_group'] = False
        return context


class GroupWordsView(BaseGroupView):
    model = Word
    context_object_name = 'words'
    paginate_by = 25

    def dispatch(self, request, *args, **kwargs):
        group_id = self.kwargs.get('group_id')
        group = get_object_or_404(WordGroup, id=group_id)

        if request.user != group.user and request.user not in group.uses_users.all():
            return HttpResponseRedirect(reverse('groups'))
        return super().dispatch(request, *args, **kwargs)

    def get_context_data(self, **kwargs):
        context = self.get_common_context(**kwargs)

        group_id = self.kwargs.get('group_id')
        group = get_object_or_404(WordGroup, id=group_id)
        is_my_group = group.user == self.request.user

        context.update({
            'is_my_group': is_my_group,
            'is_uses': not is_my_group and group.uses_users.filter(id=self.request.user.id).exists(),
            'title': "Groups",
            'title1': f"{group.name} Words ({group.user.username})" if not is_my_group else f"{group.name} Words",
            'is_main': group.is_main,
            'group_id': group_id,
            'group_': group,
            'is_group': True,
            'words_f_g': True,
        })

        if context['is_uses']:
            all_words = Word.objects.filter(user=self.request.user)

            user_word_titles = [word.word.lower() for word in all_words]

            group_words = group.words.all()

            for word in group_words:
                word.is_saved = word.word.lower() in user_word_titles

            context['words'] = group_words
        return context

    def get_queryset(self):
        group_id = self.kwargs.get('group_id')
        group = get_object_or_404(WordGroup, id=group_id)
        return group.words.all()


class CreateGroupView(LoginRequiredMixin, View):
    def get(self, request, *args, **kwargs):

        len_groups = WordGroup.objects.filter(user=request.user).count()

        if request.user.username == 'grouper' and len_groups >= 101:
            return redirect('groups')

        if len_groups >= (MAX_GROUP_COUNT + 1) and request.user.username != 'grouper':
            return redirect('groups')
        
        form = GroupForm()
        return render(request, 'med/create_group.html', {'form': form})

    def post(self, request, *args, **kwargs):
        form = GroupForm(request.POST)
        if form.is_valid():
            group_name = form.cleaned_data['name']
            if WordGroup.objects.filter(name=group_name, user=request.user).exists():
                form.add_error('name', "Group with this name already exists")
                return render(request, 'med/create_group.html', {'form': form})

            group = form.save(commit=False)
            group.user = request.user
            group.save()
            return redirect('groups')

        return render(request, 'med/create_group.html', {'form': form})


def activate_account(request, uidb64, token):
    try:
        uid = urlsafe_base64_decode(uidb64).decode()
        user = User.objects.get(pk=uid)
    except (TypeError, ValueError, OverflowError, User.DoesNotExist):
        user = None

    if user is not None and default_token_generator.check_token(user, token):
        user.is_active = True
        user.save()
        messages.success(request, "Your account has been activated successfully!")
        return redirect('login')
    else:
        messages.error(request, "Activation link is invalid!")
        return redirect('register')


class RegisterUser(CreateView):
    form_class = RegisterUserForm
    template_name = 'med/register.html'
    success_url = '/login'
    extra_context = {'title': 'Register'}

    @transaction.atomic
    def form_valid(self, form):
        user = form.save(commit=False)
        if User.objects.filter(email=user.email).exists():
            form.add_error('email', "User with this email already exists")
            return self.form_invalid(form)
        
        # user.is_active = False 
        user.save()

        # send_activation_email.delay(user.id)

        # return render(self.request, 'med/activation_email_sent.html')
        login(self.request, user)
        return redirect('profile', user_name=user.username)

class LoginUser(LoginView):
    form_class =  LoginUserForm
    template_name = 'med/login.html'
    extra_context = {'title': 'Log in'}

    def get_success_url(self):
        return reverse_lazy('profile', kwargs={'user_name': self.request.user.username})


class ProfileView(LoginRequiredMixin, View):
    def get(self, request, user_name, **kwargs):
        def get_user_data(user):
            user_profile, created = UserProfile.objects.get_or_create(user=user)
            is_favorite = user_profile.what_type_show == 'fav'

            recent_words = Word.objects.filter(
                user=user, 
                is_favourite=is_favorite
            )[:user_profile.words_num_in_prof] if is_favorite else Word.objects.filter(user=user)[:user_profile.words_num_in_prof]

            words = Word.objects.filter(user=user)
            word_count = words.count()
            group_count = max((WordGroup.objects.filter(user=user).count() +
                            WordGroup.objects.filter(uses_users=self.request.user).count()) - 1, 0)

            friends = User.objects.filter(
                Q(friendship_requests_sent__receiver=user, friendship_requests_sent__status='accepted') |
                Q(friendship_requests_received__sender=user, friendship_requests_received__status='accepted')
            ).distinct()

            word_stats = words.values('word_type').annotate(count=Count('id'))
            word_type_data = json.dumps([
                {
                    'name': word_type['word_type'].capitalize(),
                    'y': word_type['count']
                }
                for word_type in word_stats
            ])

            n_days = 7
            today = now().date()
            start_date = today - timedelta(days=n_days)

            daily_word_count = Word.objects.filter(
                user=user,
                time_create__gte=start_date
            ).annotate(
                day=TruncDay('time_create')
            ).values('day').annotate(
                count=Count('id')
            ).order_by('day')

            daily_data = {
                entry['day'].date().isoformat(): entry['count']
                for entry in daily_word_count
            }
            daily_chart_data = json.dumps({
                'categories': [(start_date + timedelta(days=i)).strftime('%Y-%m-%d') for i in range(n_days + 1)],
                'data': [daily_data.get((start_date + timedelta(days=i)).strftime('%Y-%m-%d'), 0) for i in range(n_days + 1)],
            })

            if user_profile.chenged_order:
                achievements_order = user_profile.achicment_order.strip('[]').replace('"', '').split(",")

                order = [int(i) for i in achievements_order]

                achievements = UserAchievement.objects.filter(
                    user=request.user,
                    id__in=achievements_order
                ).order_by(Case(*[When(id=id, then=pos) for pos, id in enumerate(order)]))
            else:
                achievements = UserAchievement.objects.filter(user=request.user)[:5]
            
            
            # process_special_achivments(user)
            # process_interaction_achivments(user)

            streak_data = UserProfile.calculate_streak(user)

            current_streak, max_streak, current_start, current_end, longest_start, longest_end, ff = streak_data

            current_streak_range = UserProfile.format_date_range(current_start, current_end)
            longest_streak_range = UserProfile.format_date_range(longest_start, longest_end)

            user_in_top = Top.objects.filter(user=user, place__in=[1, 2, 3]) or None

            return {
                'user_profile': user_profile,
                'recent_words': recent_words,
                'word_count': word_count,
                'group_count': group_count,
                'friends': friends,
                'friend_count': friends.count(),
                'is_favorite': is_favorite,
                'word_type_data': word_type_data,
                'daily_chart_data': daily_chart_data,
                'order': user_profile.charts_order.split(','),
                'achievements': achievements,
                'current_streak': current_streak,
                'current_streak_range': current_streak_range,
                'longest_streak': max_streak,
                'longest_streak_range': longest_streak_range,
                'ff': ff,
                'user_in_top': user_in_top,
            }

        profile_user = get_object_or_404(User, username=user_name)
        is_my_profile = profile_user == request.user

        profile_data = get_user_data(profile_user)

        is_requests_in = True if Friendship.objects.filter(receiver=request.user, sender=profile_user, status='pending').values_list('sender', flat=True) else False
        is_requests_out = True if Friendship.objects.filter(sender=request.user, receiver=profile_user, status='pending').values_list('receiver', flat=True) else False

        if not is_my_profile:
            profile_data['is_friends'] = request.user in profile_data['friends']
        else:
            profile_data['friend_requests'] = Friendship.objects.filter(receiver=request.user, status='pending')

        friendships = Friendship.objects.filter(
            Q(sender=profile_user, status='accepted') |
            Q(receiver=profile_user, status='accepted')
        )

        if not is_my_profile and profile_data['is_friends']:
            friendship = friendships.filter(sender=profile_user, receiver=request.user).first() or friendships.filter(sender=request.user, receiver=profile_user).first()
            profile_user.friendship_id = friendship.id if friendship else None

        # For test adding new notifications
        # Notification.objects.create(
        #     user=request.user,
        #    message='This is a test notification',
        #     is_read=False,
        # )

        return render(request, 'med/profile.html', {
            **profile_data,
            'user': profile_user,
            'logged_user': request.user,
            'is_my_profile': is_my_profile,
            'is_profile': True,
            'is_requests_in': is_requests_in,
            'is_requests_out': is_requests_out,
        })
    

class SelectGroupView(View):
    def get(self, request):
        word_ids = request.GET.getlist('word_ids')
        if not word_ids:
            return redirect('words', user_name=request.user.username)
        words = Word.objects.filter(id__in=word_ids)
        groups = WordGroup.objects.filter(user=request.user, is_main=False)
        return render(request, 'med/select_group.html', {
            'words': words,
            'word_ids': word_ids,
            'groups': groups,
        })

    def post(self, request, *args, **kwargs):
        word_ids = request.POST.getlist('word_ids')
        group_id = request.POST.get('group')

        if group_id:
            group = get_object_or_404(WordGroup, id=group_id, user=request.user)
        else:
            redirect('words', user_name=request.user.username)
        
        group_words = group.words.all()
        words = Word.objects.filter(id__in=word_ids, user=request.user)

        for word in words:
            if word not in group_words:
                group.words.add(word)

        group.save()

        return redirect('group_words', group_id=group_id)
    

class EditProfileView(LoginRequiredMixin, View):
    def get_user_profile(self, user):
        return UserProfile.objects.get_or_create(user=user)[0]

    def get_forms(self, request, user_profile):
        return {
            'profile_form': EditProfileForm(instance=request.user),
            'words_show_form': WordsShowForm(instance=user_profile),
            'avatar_form': AvatarUpdateForm(instance=user_profile),
            'password_form': ChengePasswordForm(user=request.user),
        }

    def render_profile_page(self, request, user_profile, **kwargs):
        forms = self.get_forms(request, user_profile)
        forms.update(kwargs)
        forms['user_profile'] = user_profile 
        forms['order'] = user_profile.charts_order.split(',')
        return render(request, 'med/edit_profile.html', forms)

    def handle_delete_avatar(self, user_profile):
        if user_profile.avatar:
            try:
                os.remove(user_profile.avatar.path.replace('/media/', '/media/avatars/'))
            except FileNotFoundError:
                pass
            user_profile.avatar = None
            user_profile.save()

    def handle_cropped_avatar(self, request, user_profile):
        avatar_data = request.POST.get('cropped_avatar')
        if avatar_data:
            try:
                format, imgstr = avatar_data.split(';base64,')
                extension = format.split('/')[-1]
                avatar_file = ContentFile(base64.b64decode(imgstr))
                new_name = f"{request.user.username}_{now().strftime('%Y%m%d%H%M')}_{request.user.id}.{extension}"

                if user_profile.avatar:
                    os.remove(user_profile.avatar.path.replace('/media/', '/media/avatars/'))

                user_profile.avatar.save(new_name, avatar_file)
                return True
            except Exception as e:
                return False
        return False

    def get(self, request, *args, **kwargs):
        user_profile = self.get_user_profile(request.user)
        return self.render_profile_page(request, user_profile)

    def post(self, request, *args, **kwargs):
        user_profile = self.get_user_profile(request.user)
        is_ajax = request.headers.get('X-Requested-With') == 'XMLHttpRequest'

        if 'delete_avatar' in request.POST:
            self.handle_delete_avatar(user_profile)
            if is_ajax:
                return JsonResponse({'status': 'success', 'message': 'Avatar deleted'})
            return redirect('profile', user_name=request.user.username)

        if 'update_profile' in request.POST:
            profile_form = EditProfileForm(request.POST, instance=request.user)
            if profile_form.is_valid():
                profile_form.save()
                if is_ajax:
                    return JsonResponse({'status': 'success', 'message': 'Profile updated'})
                return redirect('profile', user_name=request.user.username)
            if is_ajax:
                return JsonResponse({'status': 'error', 'errors': profile_form.errors})
            return self.render_profile_page(request, user_profile, profile_form=profile_form)

        if 'update_words_show' in request.POST:
            words_show_form = WordsShowForm(request.POST, instance=user_profile)
            charts_order = request.POST.get('word_stat_order')
            if charts_order:
                user_profile.charts_order = charts_order
                user_profile.save()

            pie_visible = request.POST.get('pie-visible') == 'true'
            bar_visible = request.POST.get('bar-visible') == 'true'
            line_visible = request.POST.get('time-visible') == 'true'

            if pie_visible != user_profile.show_pie_chart:
                user_profile.show_pie_chart = pie_visible
            if bar_visible != user_profile.show_bar_chart:
                user_profile.show_bar_chart = bar_visible
            if line_visible != user_profile.show_line_chart:
                user_profile.show_line_chart = line_visible

            if words_show_form.is_valid():
                words_show_form.save()
                user_profile.save()
                if is_ajax:
                    return JsonResponse({'status': 'success', 'message': 'Words settings updated'})
                return redirect('profile', user_name=request.user.username)
            if is_ajax:
                return JsonResponse({'status': 'error', 'errors': words_show_form.errors})
            return self.render_profile_page(request, user_profile, words_show_form=words_show_form)

        if 'cropped_avatar' in request.POST:
            success = self.handle_cropped_avatar(request, user_profile)
            if is_ajax:
                return JsonResponse({
                    'status': 'success' if success else 'error',
                    'message': 'Avatar updated' if success else 'Failed to update avatar'
                })
            return redirect('profile', user_name=request.user.username)

        if 'change_password' in request.POST:
            password_form = ChengePasswordForm(user=request.user, data=request.POST)
            if password_form.is_valid():
                password_form.save()
                update_session_auth_hash(request, request.user)
                if is_ajax:
                    return JsonResponse({'status': 'success', 'message': 'Password updated'})
                return redirect('profile', user_name=request.user.username)
            if is_ajax:
                return JsonResponse({'status': 'error', 'errors': password_form.errors})
            return self.render_profile_page(request, user_profile, password_form=password_form)

        if is_ajax:
            return JsonResponse({'status': 'error', 'message': 'Invalid request'})
        return self.render_profile_page(request, user_profile)
        

def logout_user(request):
    logout(request)
    return redirect('login')

def make_favourite(request, word_id):
    try:
        word = get_object_or_404(Word, id=word_id, user=request.user)
        word.is_favourite = not word.is_favourite
        word.save()
        return redirect('words', user_name=request.user.username)
    except:
        return redirect('login')

def check_username(request):
    if request.method == 'POST':
        data = json.loads(request.body)
        username = data.get('username', None)

        if username and User.objects.filter(username=username).exists():
            return JsonResponse({'error': 'This username is already taken.'}, status=200)
        return JsonResponse({'error': ''}, status=200)

def user_search(request):
    query = request.GET.get('q')
    users = User.objects.filter(username__icontains=query) if query else []
    friends = User.objects.filter(
        Q(friendship_requests_sent__receiver=request.user, friendship_requests_sent__status='accepted') |
        Q(friendship_requests_received__sender=request.user, friendship_requests_received__status='accepted')
    ).distinct()

    friendships = Friendship.objects.filter(
        (Q(sender=request.user) & Q(receiver__in=friends)) | (Q(receiver=request.user) & Q(sender__in=friends))
    )

    friend_requests_in = Friendship.objects.filter(receiver=request.user, status='pending').values_list('sender_id', flat=True)
    friend_requests_out = Friendship.objects.filter(sender=request.user, status='pending').values_list('receiver_id', flat=True)

    for user in users:
        friendship = friendships.filter(sender=user, receiver=request.user).first() or friendships.filter(sender=request.user, receiver=user).first()
        user.friendship_id = friendship.id if friendship else None

    paginator = Paginator(users, 15)
    page_number = request.GET.get('page')
    paginated_users = paginator.get_page(page_number)
    
    for user in users:
        user.img = user.user_profile.get_avatar_url()

    return render(request, 'med/user_search.html', {
        'users': paginated_users,
        'query': query,
        'friends': friends,
        'friend_requests_in': friend_requests_in,
        'friend_requests_out': friend_requests_out,
        'is_search': True,
        'paginator': paginator,
    })

@login_required
def send_friend_request(request, username):
    receiver = get_object_or_404(User, username=username)

    if receiver == request.user:
        return redirect('profile', user_name=request.user.username)
    
    friendship, created = Friendship.objects.get_or_create(sender=request.user, receiver=receiver)

    Notification.objects.create(
        user=receiver,
        message=f"{request.user.username} sent you a friend request",
        is_read=False,
    )

    return redirect('profile', user_name=request.user.username)

@login_required
def respond_to_friend_request(request, friendship_id=None, user1_id=None, user2_id=None, response=None):
    if friendship_id:
        friendship = get_object_or_404(Friendship, id=friendship_id)
    elif user1_id and user2_id:
        user1 = get_object_or_404(User, id=user1_id)
        user2 = get_object_or_404(User, id=user2_id)
        friendship = Friendship.objects.filter(
            (Q(sender=user1) & Q(receiver=user2)) | 
            (Q(sender=user2) & Q(receiver=user1)), 
            status='pending'
        ).first()
        if not friendship:
            return redirect('friends_list', user_name=request.user.username)
    else:
        return redirect('friends_list', user_name=request.user.username)

    if friendship.receiver == request.user and response == 'accept':
        friendship.status = 'accepted'
        friendship.save()
    elif (friendship.receiver == request.user or friendship.sender == request.user) and response == 'reject':
        friendship.delete()
    else:
        pass

    return redirect('friends_list', user_name=request.user.username)

def delete_friend(request, friendship_id):
    friendship = get_object_or_404(Friendship, id=friendship_id)

    if friendship.sender == request.user or friendship.receiver == request.user:
        friendship.delete()

    return redirect('friends_list', user_name=request.user.username)

@login_required
def friends_list_view(request, user_name):
    user = get_object_or_404(User, username=user_name)
    is_my_friends = user == request.user

    friendships = Friendship.objects.filter(
        Q(sender=user, status='accepted') |
        Q(receiver=user, status='accepted')
    ).select_related('sender', 'receiver')

    friends = [
        friendship.sender if friendship.receiver == user else friendship.receiver
        for friendship in friendships
    ]

    friend_requests_in = Friendship.objects.filter(receiver=user, status='pending').select_related('sender')
    friend_requests_out = Friendship.objects.filter(sender=user, status='pending').select_related('receiver')

    for friend in friends:
        friendship = friendships.filter(sender=friend, receiver=request.user).first() or friendships.filter(sender=request.user, receiver=friend).first()
        friend.friendship_id = friendship.id if friendship else None

    return render(request, 'med/friends_list.html', {
        'user': user,
        'friends': friends,
        'friend_requests_in': friend_requests_in,
        "in_count": friend_requests_in.count(),
        'friend_requests_out': friend_requests_out,
        'is_my_friends': is_my_friends,
    })

def practice_view(request):
    return render(request, 'med/practice.html', {'cards': practice_cards.values()})

def reading_view(request):
    texts = ReadingText.objects.all()
    
    word_stats = texts.aggregate(min_words=Min('word_count'), max_words=Max('word_count'))
    min_words = word_stats['min_words'] or 0 
    max_words = word_stats['max_words'] or 0

    if request.method == 'GET':
        if 'words_min' not in request.GET or 'words_max' not in request.GET:
            query_params = request.GET.copy()
            if 'words_min' not in request.GET:
                query_params['words_min'] = min_words
            if 'words_max' not in request.GET:
                query_params['words_max'] = max_words
            return HttpResponseRedirect(f"{request.path}?{query_params.urlencode()}")

    level = request.GET.get('level')
    if level:
        texts = texts.filter(eng_level=level)

    words_min = request.GET.get('words_min', min_words)
    if words_min:
        texts = texts.filter(word_count__gte=int(words_min))

    words_max = request.GET.get('words_max', max_words)
    if words_max:
        texts = texts.filter(word_count__lte=int(words_max))

    paginator = Paginator(texts, 25)
    page_number = request.GET.get('page')
    paginated_texts = paginator.get_page(page_number)

    levels = ReadingText.ENG_LEVEL_CHOICES

    return render(request, 'med/reading.html', {
        'texts': paginated_texts,
        'paginator': paginator,
        'page_obj': paginated_texts,
        'levels': levels,
        'min_words': min_words,
        'max_words': max_words,
    })

def parct_groups_view(request):
    groups = CommunityGroup.objects.filter(state='added').select_related('group')

    for group in groups:
        group.words_count = group.group.words.count()
    
    groups = sorted(groups, key=lambda x: x.words_count)

    paginator = Paginator(groups, 25)
    page_number = request.GET.get('page')
    paginated_groups = paginator.get_page(page_number)
    page_obj = paginator.get_page(page_number)

    return render(request, 'med/practice_groups.html', {'groups': paginated_groups, 'paginator': paginator, 'page_obj': page_obj})

def split_content_by_phrases(content, translations):
    words = content.split()
    result = []
    i = 0
    lower_translations = {k.lower(): v for k, v in translations.items()}

    def clean_word(word):
        return word.strip(string.punctuation).lower()

    while i < len(words):
        match = None
        for j in range(len(words), i, -1):
            phrase = " ".join(words[i:j])
            cleaned_phrase = clean_word(phrase)
            if cleaned_phrase in lower_translations:
                match = (phrase, lower_translations[cleaned_phrase])
                i = j
                break
        if match:
            result.append(match)
        else:
            word = words[i]
            result.append((word, None))
            i += 1
    return result

def reading_text_view(request, text_id):
    text = get_object_or_404(ReadingText, id=text_id)
    base_url = None
    if text.is_auth_a:
        url = text.auth
        parsed_url = urlparse(url)
        base_url = f"{parsed_url.scheme}://{parsed_url.netloc}/"

    content_phrases = []
    for paragraph in text.content.splitlines():
        phrases = split_content_by_phrases(paragraph, text.words_with_translations)
        content_phrases.append(phrases)

    paginator = Paginator(content_phrases, 5)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    user_profile = UserProfile.objects.get(user=request.user)
    user_profile.text_read += 1
    user_profile.save()

    return render(request, 'med/read_text.html', {
        'text': text,
        'base_url': base_url,
        'page_obj': page_obj,
        'paginator': paginator,
    })

def word_detail_view(request, word, text_id):
    text = get_object_or_404(ReadingText, content__icontains=word, id=text_id)
    translation = text.words_with_translations.get(word.lower(), 'The translation is not found')

    sentences = re.split(r'(?<=[.!?])[\s\n]+', text.content)
    example_sentence = next((sentence for sentence in sentences if word in sentence), 'The example sentence is not found')

    return render(request, 'med/word_detail.html', {'word': word, 'translation': translation, 'example': example_sentence, 'text_id': text_id})

def word_couner(text):
    words = text.split()
    return len(words)

def time_counter(words):
    return round(words / 60)

def text_add_view(request):
    form = TextForm()

    if request.method == 'POST':
        form = TextForm(request.POST)
        word_count = word_couner(form.data['content'])
        time_count = time_counter(word_count)
        if form.is_valid():
            form.instance.word_count = word_count
            form.instance.time_to_read = time_count
            form.save()
            return redirect('practice_reading')
        
    return render(request, 'med/add_text.html', {'form': form})


class EditTextView(LoginRequiredMixin, View):
    def get(self, request, text_id, *args, **kwargs):
        text = get_object_or_404(ReadingText, id=text_id)
        form = TextForm(instance=text)
        return render(request, 'med/edit_text.html', {'form': form, 'text': text})

    def post(self, request, text_id, *args, **kwargs):
        text = get_object_or_404(ReadingText, id=text_id)
        form = TextForm(request.POST, instance=text)
        word_count = word_couner(form.data['content'])
        time_count = time_counter(word_count)
        if form.is_valid():
            form.instance.word_count = word_count
            form.instance.time_to_read = time_count
            form.save()
            return redirect('reading_text', text_id=text_id)
        return render(request, 'med/edit_text.html', {'form': form, 'text': text})


class PracticeGroupWordsListView(LoginRequiredMixin, ListView):
    model = Word
    template_name = 'med/group_words.html'
    context_object_name = 'words'
    paginate_by = 25

    def get_queryset(self):
        group_id = self.kwargs.get('group_id')
        group = get_object_or_404(WordGroup, id=group_id)

        user_word_titles = [word.word.lower() for word in Word.objects.filter(user=self.request.user)]

        words = group.words.all()
        for word in words:
                word.is_saved = word.word.lower() in user_word_titles

        return words
    
    def get_context_data(self, **kwargs):
        process_interaction_achivments(self.request.user)

        context = super().get_context_data(**kwargs)
        group_id = self.kwargs.get('group_id')
        group = get_object_or_404(WordGroup, id=group_id)
        context['group'] = group
        context['is_usses'] = group.uses_users.filter(id=self.request.user.id).exists()
        if CommunityGroup.objects.filter(group=context['group'], state='added').exists():
            context['is_community'] = True
        else:
            context['is_community'] = False
        return context
    
def add_as_uses(request, group_id):
    group = get_object_or_404(WordGroup, id=group_id)
    group.uses_users.add(request.user)
    return redirect('group_words_practice', group_id=group_id)

def leave_group(request, group_id, fp):

    if fp.lower() in ['true', '1', 'yes']:
        is_fp = True
    elif fp.lower() in ['false', '0', 'no']:
        is_fp = False

    group = get_object_or_404(WordGroup, id=group_id)
    group.uses_users.remove(request.user)
    if is_fp:
        return redirect('group_words_practice', group_id=group_id)
    return redirect('groups')


def save_word(request, group_id, word_id):
    word = get_object_or_404(Word, id=word_id)
    existing_word = Word.objects.filter(user=request.user, id=word.id).first()

    if not existing_word:
        Word.objects.create(user=request.user, word=word.word, translation=word.translation,
                            word_type=word.word_type, example=word.example, is_favourite=False)

        new_word = Word.objects.get(word=word.word, user=request.user)

        add_to_main_group(request, new_word)

    if group_id:
        return redirect('group_words_practice', group_id=group_id)
    else:
        return redirect('words', user_name=request.user.username)

def save_group_words(request, group_id):
    user = request.user
    group = get_object_or_404(WordGroup, id=group_id)
    words = group.words.all()
    word_ids = [word.id for word in words]
    group_name = f"All {user.username}'s "
    all_my_words = Word.objects.filter(user=user)

    new_group = WordGroup.objects.create(
        name=f"{group.name}",
        user=user
    )

    for word_id in word_ids:
        original_word = get_object_or_404(Word, id=word_id)

        if not all_my_words.filter(word=original_word.word).exists():
            copied_word = Word.objects.create(
                word=original_word.word,
                translation=original_word.translation,
                example=original_word.example,
                word_type=original_word.word_type,
                user=user,
            )
            
            # all words groop
            group1, created = WordGroup.objects.get_or_create(
                name=group_name,
                is_main=True,
                user=user
            )

            group1.words.add(copied_word)

            new_group.words.add(copied_word)

        else:
            word = all_my_words.get(word=original_word.word)
            new_group.words.add(word)
    
    group.uses_users.remove(user)
    return redirect('groups')

def upload_file(request):
    if request.method == 'POST' and request.FILES.get('file'):
        group_name = f"All {request.user.username}'s "
        group, created = WordGroup.objects.get_or_create(
            name=group_name,
            is_main=True,
            user=request.user
        )
        uploaded_file = request.FILES['file']
        file_name = uploaded_file.name.lower()

        try:
            if file_name.endswith('.xlsx') or file_name.endswith('.xls'):
                wb = openpyxl.load_workbook(uploaded_file)
                sheet = wb[wb.sheetnames[0]]
                words = []
                for row in sheet.iter_rows(values_only=True):
                    if len(row) == 3:
                        words.append(row)
                wb.close()

                if words[0] == ('Word', 'Translation', 'Example'):
                    for word in words[1:]:
                        word = Word.objects.create(
                            word=word[0],
                            translation=word[1],
                            example=word[2],
                            user=request.user,
                        )
                        group.words.add(word)
                    return redirect('words', user_name=request.user.username)
                else:
                    messages.error(request, "Excel file format is incorrect. First row should be 'Word', 'Translation', 'Example'.")
                    return render(request, 'med/words_ff.html')

            elif file_name.endswith('.txt'):
                content = uploaded_file.read().decode('utf-8')
                try:
                    json_data = json.loads(content)
                    if isinstance(json_data, list):
                        for item in json_data:
                            if all(k in item for k in ('word', 'translation', 'example')):
                                word_type = item.get('word_type', 'other') 
                                
                                word = Word.objects.create(
                                    word=item['word'],
                                    translation=item['translation'],
                                    example=item['example'],
                                    word_type=word_type, 
                                    user=request.user,
                                )
                                group.words.add(word)
                        return redirect('words', user_name=request.user.username)
                    else:
                        messages.error(request, "JSON file should contain a list of objects with keys: 'word', 'translation', 'example'.")
                        return render(request, 'med/words_ff.html')
                except json.JSONDecodeError:
                    messages.error(request, "Invalid JSON format in text file.")
                    return render(request, 'med/words_ff.html')

            else:
                messages.error(request, "Unsupported file format. Please upload .xlsx, .xls, or .txt file.")
                return render(request, 'med/words_ff.html')

        except Exception as e:
            messages.error(request, f"An error occurred while processing the file: {e}")
            return render(request, 'med/words_ff.html')

    return render(request, 'med/words_ff.html')


def download_file(request, file):
    base_dir = 'media'
    file_path = os.path.join(base_dir, file)
    
    if not os.path.exists(file_path):
        raise Http404("File does not exist")

    response = FileResponse(open(file_path, 'rb'), as_attachment=True)
    response['Content-Disposition'] = 'attachment; filename=' + file
    return response

def save_all_words_as_json(request):
    user = request.user
    words = Word.objects.filter(user=user)
    data = []

    for word in words:
        data.append({
            'word': word.word,
            'translation': word.translation,
            'example': word.example,
            'word_type': word.word_type,
        })
    
    file_name = f"{user.username}_words.json"
    file_path = os.path.join('media', file_name)

    with open(file_path, 'w') as file:
        json.dump(data, file, indent=4)

    return redirect('download_file', file=file_name)

def find_word_type(request):
    if request.method == 'POST':
        data = json.loads(request.body)
        word = data.get('word', None)
        if word:
            api_url = f"https://api.dictionaryapi.dev/api/v2/entries/en/{word}"
            response = requests.get(api_url)

            if response.status_code == 200:
                word_type = response.json()[0]['meanings'][0]['partOfSpeech']
                try:
                    all_types = set()
                    for meaning in response.json()[0]['meanings']:
                        all_types.add(meaning['partOfSpeech'])
                    all_types.remove(word_type)
                    all_types = list(all_types)

                    if len(all_types) > 1:
                        return JsonResponse({'word_type': word_type, 'all_types': all_types}, status=200)
                    else:
                        return JsonResponse({'word_type': word_type}, status=200)
                except:
                    return JsonResponse({'word_type': word_type}, status=200)
            else:
                return JsonResponse({'word_type': 'other'}, status=200)
        
        return JsonResponse({'word_type': 'other'}, status=200)
    
    return JsonResponse({'error': 'Invalid request method'}, status=400)

def check_word(request):
    if request.method == 'POST':
        data = json.loads(request.body)
        word = data.get('word', None)
        if word:
            if Word.objects.filter(word=word, user=request.user).exists():
                return JsonResponse({'error': 'This word already exists in your dictionary.'}, status=200)
            return JsonResponse({'error': ''}, status=200)
        return JsonResponse({'error': 'Invalid word'}, status=200)

def process_achievements(user, achievement_type, thresholds):
    if achievement_type == '6':
        processed_signals.clear()

    if user.pk in processed_signals:
        return

    processed_signals[user.pk] = True

    achievements = Achievement.objects.filter(ach_type=achievement_type)
    user_profile = UserProfile.objects.get(user=user)

    current_user_achievements = UserAchievement.objects.filter(user=user, achievement__ach_type=achievement_type)
    current_levels = {ua.achievement.level for ua in current_user_achievements}
    if achievement_type == '1':
        item_count = Word.objects.filter(user=user).count()
    elif achievement_type == '2':
        item_count = WordGroup.objects.filter(user=user, is_main=False).count()
    elif achievement_type == '3':
        item_count = Friendship.objects.filter(Q(sender=user, status='accepted') | Q(receiver=user, status='accepted')).count()
    elif achievement_type == '4':
        item_count = (user_profile.text_read, user_profile.words_added_from_text)
    elif achievement_type == '6':
        item_count = Word.objects.filter(user=user).exclude(example='').count()
    else:
        item_count = 0

    for i, ach in enumerate(achievements):
        threshold = thresholds[i] if i < len(thresholds) else thresholds[-1]

        if achievement_type in ['1', '2', '3', '6']:
            if achievement_type == '2' and current_levels == {2}:
                groups_with_more_5_words = WordGroup.objects.filter(user=user, is_main=False).annotate(
                    word_count=Count('words')
                ).filter(word_count__gte=5).count()
                if groups_with_more_5_words >= 5 and ach.level not in current_levels:
                    if any(existing_level > ach.level for existing_level in current_levels):
                        continue

                    UserAchievement.objects.filter(
                        user=user,
                        achievement__ach_type=achievement_type,
                        achievement__level__lt=ach.level
                    ).delete()

                    UserAchievement.objects.create(user=user, achievement=ach)

                    current_levels.add(ach.level)
                
                continue

            if item_count >= threshold and ach.level not in current_levels:
                if any(existing_level > ach.level for existing_level in current_levels):
                    continue

                UserAchievement.objects.filter(
                    user=user,
                    achievement__ach_type=achievement_type,
                    achievement__level__lt=ach.level
                ).delete()

                UserAchievement.objects.create(user=user, achievement=ach)

                current_levels.add(ach.level)
        elif achievement_type == '4':
            if item_count[0] >= threshold[0] and item_count[1] >= threshold[1] and ach.level not in current_levels:
                if any(existing_level > ach.level for existing_level in current_levels):
                    continue

                UserAchievement.objects.filter(
                    user=user,
                    achievement__level__lt=ach.level
                ).delete()

                UserAchievement.objects.create(user=user, achievement=ach)

                current_levels.add(ach.level)

def process_words_achivments(user, thresholds):
    if user.pk in process_wrods_signals:
        return
    
    process_wrods_signals[user.pk] = True
    
    achievement_type = ['1', '6']

    for ach_type in achievement_type:
        process_achievements(user, ach_type, thresholds)

SITE_LAUNCH_DATE = datetime(2025, 5, 15)

def process_special_achivments(user):
    user_special_achivments = UserAchievement.objects.filter(user=user, achievement__ach_type='7').values_list('achievement__name', flat=True)
    
    all_achievements = Achievement.objects.filter(ach_type='7')

    current_time = now().replace(tzinfo=None)
    if 'Early Bird' not in user_special_achivments and current_time <= SITE_LAUNCH_DATE + timedelta(days=30):
        early_bird_achievement = all_achievements.get(name='Early Bird')
        if early_bird_achievement:
            UserAchievement.objects.create(user=user, achievement=early_bird_achievement)

    # Marathoner
    if 'Marathoner' not in user_special_achivments:
        last_30_days = now().date() - timedelta(days=30)
        user_logins = user.logins.filter(date__gte=last_30_days).aggregate(count=Count('date', distinct=True))['count']
        if user_logins == 30:
            marathoner_achievement = all_achievements.get(name='Marathoner')
            if marathoner_achievement:
                UserAchievement.objects.create(user=user, achievement=marathoner_achievement)

    # Perfectionist
    if 'Perfectionist' not in user_special_achivments:
        edited_words = UserProfile.objects.filter(user=user).values_list('edited_words', flat=True).first()
        if edited_words and edited_words >= 20:
            perfectionist_achievement = all_achievements.get(name='Perfectionist')
            if perfectionist_achievement:
                UserAchievement.objects.create(user=user, achievement=perfectionist_achievement)

    # Gotta Catch 'Em All
    if "Gotta Catch 'Em All!" not in user_special_achivments:
        user_ach_count = UserAchievement.objects.filter(user=user).exclude(achievement__ach_type='7').count()
        ach_count = Achievement.objects.exclude(ach_type='7').count()
        if ach_count == user_ach_count:
            gcea_acievement = all_achievements.get(name="Gotta Catch 'Em All!")
            if gcea_acievement:
                UserAchievement.objects.create(user=user, achievement=gcea_acievement)

def process_interaction_achivments(user):
    user_interaction_achivments = UserAchievement.objects.filter(user=user, achievement__ach_type='5').values_list('achievement__name', flat=True)
    # Friendly learner
    if not user_interaction_achivments:
        user_groups = WordGroup.objects.filter(uses_users=user).count()
        if user_groups >= 5:
            UserAchievement.objects.create(user=user, achievement=Achievement.objects.get(ach_type='5', level=1))

    # Social Butterfly
    fl_ach = Achievement.objects.get(ach_type='5', level=1)
    if fl_ach.name in user_interaction_achivments:
        shered_groups = CommunityGroup.objects.filter(state='added', group__user=user).count()
        if shered_groups >= 10:
            UserAchievement.objects.create(user=user, achievement=Achievement.objects.get(ach_type='5', level=2))
            UserAchievement.objects.filter(user=user, achievement=fl_ach).delete()

@receiver(post_save, sender=Word)
def update_achievements_words(sender, instance, **kwargs):
    thresholds = [10, 50, 100]
    # process_words_achivments(instance.user, thresholds)
    # process_special_achivments(instance.user)
    # process_interaction_achivments(instance.user)

@receiver(post_save, sender=WordGroup)
def update_achievements_on_group_words_change(sender, instance, **kwargs):
    thresholds = [1, 5, 10]
    # process_achievements(instance.user, achievement_type='2', thresholds=thresholds)
    # process_special_achivments(instance.user)
    # process_interaction_achivments(instance.user)

@receiver(m2m_changed, sender=WordGroup.words.through)
def update_achievements_on_group_words_change(sender, instance, action, **kwargs):
    if action in ['post_add', 'post_remove', 'post_clear']:
        thresholds = [1, 5, 10]
        # process_achievements(instance.user, achievement_type='2', thresholds=thresholds)
        # process_special_achivments(instance.user)
        # process_interaction_achivments(instance.user)   

@receiver(post_save, sender=Friendship)
def update_achievements_friends(sender, instance, **kwargs):
    thresholds = [5, 20, 50]
    
    # process_achievements(instance.sender, achievement_type='3', thresholds=thresholds)
    # process_achievements(instance.receiver, achievement_type='3', thresholds=thresholds)

@receiver(post_save, sender=UserProfile)
def update_achievements_reading(sender, instance, **kwargs):
    thresholds = [(5, 1), (50, 10), (100, 20)]
    # process_achievements(instance.user, achievement_type='4', thresholds=thresholds)
    # process_special_achivments(instance.user)
    # process_interaction_achivments(instance.user)

def achievement_view(request):
    user_profile = UserProfile.objects.get(user=request.user)
    if request.method == 'POST':
        order = request.POST.get('word_stat_order')
        if order:
            user_profile.achicment_order = order
            user_profile.chenged_order = True
            user_profile.save(update_fields=["achicment_order", "chenged_order"]) 

    achivments = {}
    for ach in Achievement.objects.all():
        achivments[ach.ach_type] = achivments.get(ach.ach_type, []) + [ach]

    for ach_type in list(achivments.keys()):
        achivments[Achievement.ACH_TYPE_CHOICES[int(ach_type) - 1][1]] = achivments.pop(ach_type)

    user_achivments = UserAchievement.objects.filter(user=request.user).values_list('achievement__name', flat=True)

    biggest_level_each_type = {}

    for ach_type in Achievement.ACH_TYPE_CHOICES[:6]:
        ach = Achievement.objects.filter(
            ach_type=ach_type[0],
            userachievement__user=request.user
        ).order_by('-level').first()
        
        if ach:
            biggest_level_each_type[ach_type[1]] = ach.level
    
    for ach_type, ach_list in achivments.items():
        for ach in ach_list:
            if ach_type == 'Special':
                if ach.name not in user_achivments:
                    ach.name = "?"
                    ach.description = f"???"
                continue

            if ach.level > biggest_level_each_type.get(ach_type, 0):
                ach.name = "?"
                if ach.level == biggest_level_each_type.get(ach_type, 0) + 1:
                    ach.description = ach.requirements
                else:
                    ach.description = "???"

    if user_profile.chenged_order:
        achievements_order = user_profile.achicment_order.strip('[]').replace('"', '').split(",")

        order = [int(i) for i in achievements_order]

        profile_achievements = UserAchievement.objects.filter(
            user=request.user,
            id__in=achievements_order
        ).order_by(Case(*[When(id=id, then=pos) for pos, id in enumerate(order)]))
    else:
        profile_achievements = UserAchievement.objects.filter(user=request.user)[:5]

    prof_ach = [ach.achievement for ach in profile_achievements]

    user_ach_ach = [ach.achievement for ach in UserAchievement.objects.filter(user=request.user)]

    prof_ach_biggest_level = []
    for ach in user_ach_ach:
        ach_type = ach.ach_type

        if ach_type not in prof_ach_biggest_level or ach.level > prof_ach_biggest_level[ach_type].level:
            prof_ach_biggest_level.append(ach)

    special_ach = Achievement.objects.filter(ach_type='7')

    for ach in special_ach:
        if ach not in prof_ach_biggest_level:
            prof_ach_biggest_level.append(ach)

    return render(request, 'med/achievements.html', {
        'profile_achievements': profile_achievements,
        'prof_ach': prof_ach,
        'prof_ach_biggest_level': prof_ach_biggest_level,
        'achivments': achivments,
        'types': [ach[1] for ach in Achievement.ACH_TYPE_CHOICES],
        'user_achivments': user_achivments,
        'biggest_level_each_type': biggest_level_each_type,
    })

def add_achievement(request, ach_id):
    user_profile = UserProfile.objects.only('achicment_order', 'chenged_order').get(user=request.user)
    
    ach = get_object_or_404(Achievement, id=ach_id)
    
    user_achievement = get_object_or_404(UserAchievement, user=request.user, achievement=ach)

    achicment_order = user_profile.achicment_order.strip('[]').split(",") if user_profile.achicment_order else []

    achicment_order = [str(user_achievement.id)] + achicment_order
    achicment_order = achicment_order[:5]

    user_profile.achicment_order = ",".join(achicment_order)
    user_profile.chenged_order = True
    user_profile.save(update_fields=['achicment_order', 'chenged_order'])

    return redirect('achievement')

def send_group_request(request, group_id):
    group = get_object_or_404(WordGroup, id=group_id, user=request.user)

    CommunityGroup.objects.create(
        group=group,
        state='pending',
    )
    
    return redirect('group_words', group_id=group_id)

def approve_group_request(request, group_id):
    group = get_object_or_404(CommunityGroup, group_id=group_id)
    group.state = 'added'
    group.save()

    Notification.objects.create(
        user=group.group.user,
        message=f"Your group {group.group.name} was approved",
        is_read=False,
    )

    return redirect('practice_groups')

def reject_group_request(request, group_id):
    CommunityGroup.objects.filter(group_id=group_id).delete()
    group = get_object_or_404(WordGroup, id=group_id)

    Notification.objects.create(
        user=group.user,
        message=f"Your group {group.name} was rejected",
        is_read=False,
    )

    return redirect('practice_groups')

def pending_group_requests(request):
    if not request.user.is_staff:
        return redirect('profile', user_name=request.user.username)
    
    pending_requests = CommunityGroup.objects.filter(state='pending')

    return render(request, 'med/pending_group_requests.html', {'pending_requests': pending_requests})

font_path = os.path.join(settings.BASE_DIR, 'med/static/med/fonts/DejaVuSans.ttf')
bold_font_path = os.path.join(settings.BASE_DIR, 'med/static/med/fonts/DejaVuSans-Bold.ttf')

pdfmetrics.registerFont(TTFont('DejaVuSans', font_path))
pdfmetrics.registerFont(TTFont('DejaVuSans-Bold', bold_font_path))

def wrap_text(text, max_width, font, font_size, pdf):
    words = text.split()
    lines = []
    line = []

    for word in words:
        line.append(word)
        if pdf.stringWidth(' '.join(line), font, font_size) > max_width:
            line.pop()
            lines.append(' '.join(line))
            line = [word]

    if line:
        lines.append(' '.join(line))

    return lines

@login_required
def export_pdf(request):

    if not request.user.user_profile.is_premium:
        return redirect('soon')
    
    words = Word.objects.filter(user=request.user)
    
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{request.user.username}_dictionary.pdf"'
    
    pdf = canvas.Canvas(response, pagesize=letter)
    width, height = letter
    left_margin = 50
    right_margin = 50
    max_width = width - left_margin - right_margin
    y = height - 50

    pdf.setFont("DejaVuSans", 12)
    pdf.drawString(left_margin, y, f"Dictionary of {request.user.username}")
    y -= 30 
    pdf.setFont("DejaVuSans", 10)
    pdf.drawString(left_margin, y, "-" * 50)
    y -= 20

    for word in words:
        pdf.setFont("DejaVuSans-Bold", 10)
        pdf.drawString(left_margin, y, f"Word: {word.word}")
        y -= 15

        pdf.setFont("DejaVuSans", 10)
        translation_lines = wrap_text(f"Translation: {word.translation}", max_width, "DejaVuSans", 10, pdf)
        for line in translation_lines:
            pdf.drawString(left_margin, y, line)
            y -= 15

        example_lines = wrap_text(f"Example: {word.example}", max_width, "DejaVuSans", 10, pdf)
        for line in example_lines:
            pdf.drawString(left_margin, y, line)
            y -= 15

        y -= 10

        if y < 50:
            pdf.showPage()
            pdf.setFont("DejaVuSans", 10)
            y = height - 100

    pdf.save()
    return response

def tops_by_category(request):
    categories = Category.objects.prefetch_related('tops').all()
    context = {
        'categories': categories,
    }
    return render(request, 'med/tops_by_category.html', context)

def api_get_tops_by_category(request):
    categories = Category.objects.prefetch_related('tops').all()
    categories_data = [
        {
            'id': category.id,
            'name': category.name,
            'last_update': category.last_update,
            'tops': [
                {
                    'id': top.id,
                    'user': top.user.username,
                    'points': top.points,
                } for top in category.tops.all()
            ]
        } for category in categories
    ]

    return JsonResponse({'categories': categories_data})

@login_required
def hide_warning_message(request):
    if request.method == "POST":
        profile, created = UserProfile.objects.get_or_create(user=request.user)
        profile.hide_warning_message = True
        profile.save()
        return JsonResponse({"success": True})
    return JsonResponse({"success": False}, status=400)


class NotiListView(LoginRequiredMixin, ListView):
    model = Notification
    template_name = 'med/notification.html'
    context_object_name = 'notifications'
    paginate_by = 25

    def get_queryset(self):
        user_name = self.request.user.username
        user = get_object_or_404(User, username=user_name)

        queryset = Notification.objects.filter(user=user).order_by('-time_create')

        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        user_name = self.request.user.username
        user = get_object_or_404(User, username=user_name)

        context.update({
            'user': user,
            'title': f"{user_name}'s Notifications",
        })
        return context


@login_required
def notifications_api(request):
    user = request.user
    unread_notifications = Notification.objects.filter(user=user, is_read=False).order_by('-time_create')[:5]
    count = unread_notifications.count()
    
    notifications_list = [
        {
            "message": n.message,
            "time": n.time_create.strftime("%d.%m.%Y %H:%M")
        } for n in unread_notifications
    ]

    return JsonResponse({
        "count": count,
        "notifications": notifications_list
    })

@login_required
def mark_notification_as_read(request, notification_id):
    notification = get_object_or_404(Notification, id=notification_id, user=request.user)
    if not notification.is_read:
        notification.is_read = True
        notification.save()
    return JsonResponse({'status': 'success', 'is_read': True})

def page_not_found(request, exception):
    return render(request, 'med/404.html')

def soon_page(request):
    return render(request, 'med/soon.html')